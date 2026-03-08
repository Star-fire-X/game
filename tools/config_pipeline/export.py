#!/usr/bin/env python3

import argparse
import hashlib
import json
import sys
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Callable

from openpyxl import load_workbook


BUNDLE_TYPE = "gameplay"
SCHEMA_VERSION = 1

CHARACTER_CLASS_VALUES = {"WARRIOR", "MAGE", "TAOIST"}
SKILL_TYPE_VALUES = {"PHYSICAL", "MAGICAL", "BUFF", "DEBUFF", "HEAL"}
SKILL_TARGET_VALUES = {
    "SELF",
    "SINGLE_ENEMY",
    "SINGLE_ALLY",
    "AOE_ENEMY",
    "AOE_ALLY",
    "AOE_ALL",
}
AMULET_TYPE_VALUES = {"NONE", "HOLY", "POISON", "FIRE", "ICE"}


@dataclass(frozen=True)
class ColumnSpec:
    name: str
    value_type: str
    required: bool = False
    default: Any = None
    allow_missing_column: bool = False
    output_name: str | None = None


@dataclass(frozen=True)
class ReferenceSpec:
    source_label: str
    target_table: str
    target_field: str
    extractor: Callable[[dict[str, Any]], list[Any]]
    nullable: bool = False
    value_transform: Callable[[Any], Any] | None = None
    target_transform: Callable[[Any], Any] | None = None


@dataclass(frozen=True)
class TableSpec:
    name: str
    workbook: str
    sheet: str
    artifact_file: str
    payload_root: str
    primary_key: str
    columns: tuple[ColumnSpec, ...]
    references: tuple[ReferenceSpec, ...]
    exporter: Callable[[Path, "TableSpec"], "TableExport"]


@dataclass(frozen=True)
class TableExport:
    spec: TableSpec
    records: list[dict[str, Any]]

    def payload(self) -> dict[str, Any]:
        return {self.spec.payload_root: self.records}


def _identity(value: Any) -> Any:
    return value


def _clone_default(value: Any) -> Any:
    return json.loads(json.dumps(value))


def _normalize_bool(value):
    if isinstance(value, bool):
        return value
    if value is None:
        return False
    if isinstance(value, (int, float)):
        return value != 0

    lowered = str(value).strip().lower()
    if lowered in {"1", "true"}:
        return True
    if lowered in {"0", "false", ""}:
        return False
    raise ValueError(f"invalid boolean value: {value!r}")


def _normalize_int(value, default):
    if value is None or value == "":
        return default
    return int(value)


def _normalize_float(value, default):
    if value is None or value == "":
        return default
    return float(value)


def _normalize_str(value, default):
    if value is None:
        return default
    text = str(value).strip()
    return text if text else default


def _normalize_enum(value, default, allowed_values, field_name):
    normalized = _normalize_str(value, default)
    if normalized not in allowed_values:
        raise ValueError(f"invalid {field_name}: {normalized!r}")
    return normalized


def _normalize_array(value, default, item_type):
    if value is None or value == "":
        return list(default)

    if isinstance(value, str):
        text = value.strip()
        if not text:
            return list(default)
        if text.startswith("["):
            raw_values = json.loads(text)
        else:
            raw_values = [part.strip() for part in text.split(",")]
    elif isinstance(value, (list, tuple)):
        raw_values = list(value)
    else:
        raw_values = [value]

    normalized = []
    for item in raw_values:
        if item_type == "int":
            normalized.append(int(item))
        elif item_type == "float":
            normalized.append(float(item))
        else:
            normalized.append(str(item))
    return normalized


def _normalize_json_text(value, default, field_name):
    if value is None or value == "":
        return _clone_default(default)

    if isinstance(value, str):
        text = value.strip()
        if not text:
            return _clone_default(default)
        try:
            return json.loads(text)
        except json.JSONDecodeError as exc:
            raise ValueError(f"invalid {field_name}: {exc.msg}") from exc

    if isinstance(value, (list, dict)):
        return value
    raise ValueError(f"invalid {field_name}: expected JSON text")


def _normalize_field(value, column: ColumnSpec):
    field_type = column.value_type
    default = column.default

    if field_type == "int":
        return _normalize_int(value, default)
    if field_type == "float":
        return _normalize_float(value, default)
    if field_type == "str":
        return _normalize_str(value, default)
    if field_type == "bool":
        return _normalize_bool(value if value is not None else default)
    if field_type == "character_class":
        return _normalize_enum(value, default, CHARACTER_CLASS_VALUES, column.name)
    if field_type == "skill_type":
        return _normalize_enum(value, default, SKILL_TYPE_VALUES, column.name)
    if field_type == "skill_target":
        return _normalize_enum(value, default, SKILL_TARGET_VALUES, column.name)
    if field_type == "amulet_type":
        return _normalize_enum(value, default, AMULET_TYPE_VALUES, column.name)
    if field_type == "int_array":
        return _normalize_array(value, default, "int")
    if field_type == "json":
        return _normalize_json_text(value, default, column.name)
    raise ValueError(f"unsupported field type: {field_type}")


def _row_is_empty(values):
    return all(value is None or str(value).strip() == "" for value in values)


def _column_map(spec: TableSpec) -> dict[str, ColumnSpec]:
    return {column.name: column for column in spec.columns}


def _load_sheet_rows(source_dir: Path, spec: TableSpec):
    workbook_path = source_dir / spec.workbook
    if not workbook_path.exists():
        raise ValueError(f"{spec.name} workbook not found: {workbook_path}")

    workbook = load_workbook(workbook_path, data_only=True)
    if len(workbook.sheetnames) != 1:
        raise ValueError(
            f"{spec.name} workbook must contain exactly one sheet named {spec.sheet}"
        )
    sheet = workbook.active
    if sheet.title != spec.sheet:
        raise ValueError(
            f"{spec.name} workbook sheet mismatch: expected {spec.sheet}, found {sheet.title}"
        )

    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise ValueError(f"{spec.name} workbook is empty")

    headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
    header_map = {name: index for index, name in enumerate(headers) if name}
    _validate_column_contract(spec, header_map)
    return rows, header_map


def _validate_column_contract(spec: TableSpec, header_map: dict[str, int]):
    missing_required = []
    missing_forbidden = []
    for column in spec.columns:
        if column.name in header_map:
            continue
        if column.required:
            missing_required.append(column.name)
        elif not column.allow_missing_column:
            missing_forbidden.append(column.name)

    if missing_required:
        raise ValueError(f"missing required columns: {', '.join(sorted(missing_required))}")
    if missing_forbidden:
        raise ValueError(
            f"{spec.name} missing columns not allowed to default: "
            + ", ".join(sorted(missing_forbidden))
        )


def _read_cell(row, header_map, column: ColumnSpec):
    index = header_map.get(column.name)
    if index is None or index >= len(row):
        return None
    return row[index]


def _normalize_record(row, header_map, spec: TableSpec):
    record = {}
    for column in spec.columns:
        raw = _read_cell(row, header_map, column)
        output_name = column.output_name or column.name
        record[output_name] = _normalize_field(raw, column)
    return record


def _normalize_map_list(value, field_name, required_keys, optional_keys=None):
    optional_keys = optional_keys or {}
    if not isinstance(value, list):
        raise ValueError(f"invalid {field_name}: expected JSON array")

    normalized = []
    for item in value:
        if not isinstance(item, dict):
            raise ValueError(f"invalid {field_name}: entries must be objects")
        normalized_item = {}
        for key in required_keys:
            if key not in item:
                raise ValueError(f"invalid {field_name}: missing {key}")
            normalized_item[key] = int(item[key])
        for key, default in optional_keys.items():
            normalized_item[key] = int(item.get(key, default))
        normalized.append(normalized_item)
    return normalized


def _sort_records(records, key_name):
    records.sort(key=lambda record: record[key_name])
    return records


def _export_items(source_dir: Path, spec: TableSpec) -> TableExport:
    rows, header_map = _load_sheet_rows(source_dir, spec)

    records = []
    seen_ids = set()
    for row in rows[1:]:
        if _row_is_empty(row):
            continue
        record = _normalize_record(row, header_map, spec)
        record_id = record["id"]
        if record_id <= 0:
            raise ValueError("item id must be > 0")
        if record_id in seen_ids:
            raise ValueError(f"duplicate item id: {record_id}")
        seen_ids.add(record_id)
        records.append(record)

    return TableExport(spec, _sort_records(records, "id"))


def _export_skills(source_dir: Path, spec: TableSpec) -> TableExport:
    rows, header_map = _load_sheet_rows(source_dir, spec)

    records = []
    seen_ids = set()
    for row in rows[1:]:
        if _row_is_empty(row):
            continue
        record = _normalize_record(row, header_map, spec)
        record_id = record["id"]
        if record_id <= 0:
            raise ValueError("skill id must be > 0")
        if record_id in seen_ids:
            raise ValueError(f"duplicate skill id: {record_id}")
        seen_ids.add(record_id)
        records.append(record)

    return TableExport(spec, _sort_records(records, "id"))


def _export_maps(source_dir: Path, spec: TableSpec) -> TableExport:
    rows, header_map = _load_sheet_rows(source_dir, spec)

    records = []
    seen_map_ids = set()
    for row in rows[1:]:
        if _row_is_empty(row):
            continue

        normalized = _normalize_record(row, header_map, spec)
        map_id = normalized["map_id"]
        if map_id <= 0:
            raise ValueError("map id must be > 0")
        if map_id in seen_map_ids:
            raise ValueError(f"duplicate map_id: {map_id}")
        seen_map_ids.add(map_id)

        record = {
            "map_id": map_id,
            "is_safe_zone": normalized["is_safe_zone"],
            "is_pk_zone": normalized["is_pk_zone"],
            "no_teleport": normalized["no_teleport"],
            "no_drug": normalized["no_drug"],
            "is_dark_map": normalized["is_dark_map"],
            "no_recall": normalized["no_recall"],
            "no_random_move": normalized["no_random_move"],
            "fight_zone": normalized["fight_zone"],
            "fight3_zone": normalized["fight3_zone"],
            "min_level": normalized["min_level"],
            "max_level": normalized["max_level"],
            "mine_map": normalized["mine_map"],
            "dark_level": normalized["dark_level"],
            "exp_rate": normalized["exp_rate"],
            "drop_rate": normalized["drop_rate"],
            "home_map": normalized["home_map"],
            "home_x": normalized["home_x"],
            "home_y": normalized["home_y"],
            "pk_village_map": normalized["pk_village_map"],
            "pk_village_x": normalized["pk_village_x"],
            "pk_village_y": normalized["pk_village_y"],
            "fixes": _normalize_map_list(normalized["fixes_json"], "fixes_json", ("x", "y")),
            "safe_zones": _normalize_map_list(
                normalized["safe_zones_json"],
                "safe_zones_json",
                ("x", "y", "radius"),
            ),
            "quest_requirements": _normalize_map_list(
                normalized["quest_requirements_json"],
                "quest_requirements_json",
                ("quest_id", "quest_value"),
            ),
        }
        records.append(record)

    return TableExport(spec, _sort_records(records, "map_id"))


def _export_gates(source_dir: Path, spec: TableSpec) -> TableExport:
    rows, header_map = _load_sheet_rows(source_dir, spec)

    records = []
    seen_gate_ids = set()
    seen_source_coords = set()
    for row in rows[1:]:
        if _row_is_empty(row):
            continue

        record = _normalize_record(row, header_map, spec)
        gate_id = record["gate_id"]
        if gate_id <= 0:
            raise ValueError("gate_id must be > 0")
        if gate_id in seen_gate_ids:
            raise ValueError(f"duplicate gate_id: {gate_id}")
        seen_gate_ids.add(gate_id)

        if not record["source_map"] or not record["target_map"]:
            raise ValueError("gate source_map and target_map are required")
        if record["require_item"]:
            if record["required_item_id"] <= 0:
                raise ValueError("required_item_id must be > 0 when require_item=true")
        elif record["required_item_id"] != 0:
            raise ValueError("required_item_id must be 0 when require_item=false")

        source_coord = (record["source_map"], record["source_x"], record["source_y"])
        if source_coord in seen_source_coords:
            raise ValueError(
                "duplicate gate source coordinate: "
                f"{record['source_map']}:{record['source_x']}:{record['source_y']}"
            )
        seen_source_coords.add(source_coord)
        records.append(record)

    return TableExport(spec, _sort_records(records, "gate_id"))


def _export_drops(source_dir: Path, spec: TableSpec) -> TableExport:
    rows, header_map = _load_sheet_rows(source_dir, spec)

    tables = {}
    seen_drop_keys = set()
    for row in rows[1:]:
        if _row_is_empty(row):
            continue

        record = _normalize_record(row, header_map, spec)
        if record["monster_template_id"] <= 0:
            raise ValueError("monster_template_id must be > 0")
        if record["item_id"] <= 0:
            raise ValueError("item_id must be > 0")
        if record["drop_rate"] < 0.0 or record["drop_rate"] > 1.0:
            raise ValueError("drop_rate must be in [0,1]")
        if (
            record["min_count"] < 0
            or record["max_count"] < 0
            or record["min_count"] > record["max_count"]
        ):
            raise ValueError(
                "min_count and max_count must be non-negative with min_count <= max_count"
            )

        drop_key = (record["monster_template_id"], record["item_id"])
        if drop_key in seen_drop_keys:
            raise ValueError(
                "duplicate drop entry: "
                f"monster_template_id={record['monster_template_id']}, "
                f"item_id={record['item_id']}"
            )
        seen_drop_keys.add(drop_key)

        tables.setdefault(record["monster_template_id"], []).append(
            {
                "item_id": record["item_id"],
                "drop_rate": record["drop_rate"],
                "min_count": record["min_count"],
                "max_count": record["max_count"],
                "rarity": record["rarity"],
                "boss_bonus": record["boss_bonus"],
            }
        )

    records = []
    for monster_template_id in sorted(tables):
        items = sorted(tables[monster_template_id], key=lambda item: item["item_id"])
        records.append({"monster_template_id": monster_template_id, "items": items})
    return TableExport(spec, records)


def _export_shops(source_dir: Path, spec: TableSpec) -> TableExport:
    rows, header_map = _load_sheet_rows(source_dir, spec)

    records = []
    seen_store_ids = set()
    for row in rows[1:]:
        if _row_is_empty(row):
            continue

        record = _normalize_record(row, header_map, spec)
        store_id = record["store_id"]
        if store_id <= 0:
            raise ValueError("store_id must be > 0")
        if store_id in seen_store_ids:
            raise ValueError(f"duplicate store_id: {store_id}")
        seen_store_ids.add(store_id)

        items = record["items_json"]
        if not isinstance(items, list):
            raise ValueError("invalid items_json: expected JSON array")

        normalized_items = []
        seen_item_ids = set()
        for item in items:
            if not isinstance(item, dict):
                raise ValueError("invalid items_json: entries must be objects")

            item_id = _normalize_int(item.get("item_id"), 0)
            price = _normalize_int(item.get("price"), 0)
            stock = _normalize_int(item.get("stock"), -1)

            if item_id <= 0:
                raise ValueError("item_id must be > 0")
            if price < 0:
                raise ValueError("price must be >= 0")
            if stock != -1 and stock < 0:
                raise ValueError("stock must be -1 or >= 0")
            if item_id in seen_item_ids:
                raise ValueError(
                    f"duplicate shop item entry: store_id={store_id}, item_id={item_id}"
                )

            seen_item_ids.add(item_id)
            normalized_items.append({"item_id": item_id, "price": price, "stock": stock})

        normalized_items.sort(key=lambda item: item["item_id"])
        records.append(
            {
                "store_id": store_id,
                "name": record["name"],
                "buy_rate": record["buy_rate"],
                "sell_rate": record["sell_rate"],
                "items": normalized_items,
            }
        )

    return TableExport(spec, _sort_records(records, "store_id"))


def _export_monster_spawns(source_dir: Path, spec: TableSpec) -> TableExport:
    rows, header_map = _load_sheet_rows(source_dir, spec)

    records = []
    seen_spawn_ids = set()
    for row in rows[1:]:
        if _row_is_empty(row):
            continue

        record = _normalize_record(row, header_map, spec)
        if record["spawn_id"] <= 0:
            raise ValueError("spawn_id must be > 0")
        if record["spawn_id"] in seen_spawn_ids:
            raise ValueError(f"duplicate spawn_id: {record['spawn_id']}")
        if record["map_id"] <= 0:
            raise ValueError("map_id must be > 0")
        if record["monster_template_id"] <= 0:
            raise ValueError("monster_template_id must be > 0")
        if record["spawn_radius"] < 0:
            raise ValueError("spawn_radius must be >= 0")
        if record["patrol_radius"] < 0:
            raise ValueError("patrol_radius must be >= 0")
        if record["respawn_interval"] < 0:
            raise ValueError("respawn_interval must be >= 0")
        if record["max_count"] <= 0:
            raise ValueError("max_count must be > 0")

        seen_spawn_ids.add(record["spawn_id"])
        records.append(record)

    return TableExport(spec, _sort_records(records, "spawn_id"))


def _export_npcs(source_dir: Path, spec: TableSpec) -> TableExport:
    rows, header_map = _load_sheet_rows(source_dir, spec)

    records = []
    seen_npc_ids = set()
    for row in rows[1:]:
        if _row_is_empty(row):
            continue

        record = _normalize_record(row, header_map, spec)
        if record["npc_id"] <= 0:
            raise ValueError("npc_id must be > 0")
        if record["npc_id"] in seen_npc_ids:
            raise ValueError(f"duplicate npc_id: {record['npc_id']}")
        if record["template_id"] <= 0:
            raise ValueError("template_id must be > 0")
        if record["type"] != "MERCHANT":
            raise ValueError("type must be MERCHANT")
        if record["map_id"] <= 0:
            raise ValueError("map_id must be > 0")
        if record["x"] < 0 or record["y"] < 0:
            raise ValueError("coordinates must be >= 0")
        if record["direction"] < 0 or record["direction"] > 7:
            raise ValueError("direction must be in [0,7]")
        if record["store_id"] <= 0:
            raise ValueError("store_id must be > 0")

        seen_npc_ids.add(record["npc_id"])
        records.append(record)

    return TableExport(spec, _sort_records(records, "npc_id"))


ITEM_COLUMNS = (
    ColumnSpec("id", "int", required=True, default=0),
    ColumnSpec("name", "str", required=True, default=""),
    ColumnSpec("std_mode", "int", default=0, allow_missing_column=True),
    ColumnSpec("shape", "int", default=0, allow_missing_column=True),
    ColumnSpec("weight", "int", default=0, allow_missing_column=True),
    ColumnSpec("ani_count", "int", default=0, allow_missing_column=True),
    ColumnSpec("special_pwr", "int", default=0, allow_missing_column=True),
    ColumnSpec("item_desc", "int", default=0, allow_missing_column=True),
    ColumnSpec("looks", "int", default=0, allow_missing_column=True),
    ColumnSpec("dura_max", "int", default=0, allow_missing_column=True),
    ColumnSpec("ac", "int", default=0, allow_missing_column=True),
    ColumnSpec("mac", "int", default=0, allow_missing_column=True),
    ColumnSpec("dc", "int", default=0, allow_missing_column=True),
    ColumnSpec("mc", "int", default=0, allow_missing_column=True),
    ColumnSpec("sc", "int", default=0, allow_missing_column=True),
    ColumnSpec("need_type", "int", default=0, allow_missing_column=True),
    ColumnSpec("need_level", "int", default=0, allow_missing_column=True),
    ColumnSpec("need_class", "int", default=99, allow_missing_column=True),
    ColumnSpec("price", "int", default=0, allow_missing_column=True),
    ColumnSpec("stackable", "bool", default=False, allow_missing_column=True),
    ColumnSpec("stack_limit", "int", default=1, allow_missing_column=True),
)

SKILL_COLUMNS = (
    ColumnSpec("id", "int", required=True, default=0),
    ColumnSpec("name", "str", required=True, default=""),
    ColumnSpec("description", "str", default="", allow_missing_column=True),
    ColumnSpec("required_class", "character_class", required=True, default=None),
    ColumnSpec("required_level", "int", required=True, default=0),
    ColumnSpec("max_level", "int", default=3, allow_missing_column=True),
    ColumnSpec("train_level_req", "int_array", default=[0, 0, 0, 0], allow_missing_column=True),
    ColumnSpec("train_points_req", "int_array", default=[0, 0, 0, 0], allow_missing_column=True),
    ColumnSpec("skill_type", "skill_type", required=True, default=None),
    ColumnSpec("target_type", "skill_target", required=True, default=None),
    ColumnSpec("is_universal", "bool", default=False, allow_missing_column=True),
    ColumnSpec("is_passive", "bool", default=False, allow_missing_column=True),
    ColumnSpec("mp_cost", "int", default=0, allow_missing_column=True),
    ColumnSpec("consumes_talisman", "bool", default=False, allow_missing_column=True),
    ColumnSpec("talisman_cost", "int", default=0, allow_missing_column=True),
    ColumnSpec("required_amulet", "amulet_type", default="NONE", allow_missing_column=True),
    ColumnSpec("amulet_cost", "int", default=0, allow_missing_column=True),
    ColumnSpec("cooldown_ms", "int", default=0, allow_missing_column=True),
    ColumnSpec("cast_time_ms", "int", default=0, allow_missing_column=True),
    ColumnSpec("can_be_interrupted", "bool", default=True, allow_missing_column=True),
    ColumnSpec("range", "float", default=0.0, allow_missing_column=True),
    ColumnSpec("aoe_radius", "float", default=0.0, allow_missing_column=True),
    ColumnSpec("min_power", "int", default=0, allow_missing_column=True),
    ColumnSpec("max_power", "int", default=0, allow_missing_column=True),
    ColumnSpec("def_power", "int", default=0, allow_missing_column=True),
    ColumnSpec("def_max_power", "int", default=0, allow_missing_column=True),
    ColumnSpec("train_lv", "int", default=0, allow_missing_column=True),
    ColumnSpec("duration_ms", "int", default=0, allow_missing_column=True),
    ColumnSpec("stat_modifier", "int", default=0, allow_missing_column=True),
    ColumnSpec("dot_damage", "int", default=0, allow_missing_column=True),
    ColumnSpec("dot_interval_ms", "int", default=1000, allow_missing_column=True),
    ColumnSpec("effect_type", "int", default=0, allow_missing_column=True),
    ColumnSpec("effect_id", "int", default=0, allow_missing_column=True),
    ColumnSpec("animation_id", "str", default="", allow_missing_column=True),
    ColumnSpec("sound_id", "str", default="", allow_missing_column=True),
)

MAP_COLUMNS = (
    ColumnSpec("id", "int", required=True, default=0, output_name="map_id"),
    ColumnSpec("is_safe_zone", "bool", default=False, allow_missing_column=True),
    ColumnSpec("is_pk_zone", "bool", default=False, allow_missing_column=True),
    ColumnSpec("no_teleport", "bool", default=False, allow_missing_column=True),
    ColumnSpec("no_drug", "bool", default=False, allow_missing_column=True),
    ColumnSpec("is_dark_map", "bool", default=False, allow_missing_column=True),
    ColumnSpec("no_recall", "bool", default=False, allow_missing_column=True),
    ColumnSpec("no_random_move", "bool", default=False, allow_missing_column=True),
    ColumnSpec("fight_zone", "bool", default=False, allow_missing_column=True),
    ColumnSpec("fight3_zone", "bool", default=False, allow_missing_column=True),
    ColumnSpec("min_level", "int", default=1, allow_missing_column=True),
    ColumnSpec("max_level", "int", default=255, allow_missing_column=True),
    ColumnSpec("mine_map", "int", default=0, allow_missing_column=True),
    ColumnSpec("dark_level", "int", default=0, allow_missing_column=True),
    ColumnSpec("exp_rate", "float", default=1.0, allow_missing_column=True),
    ColumnSpec("drop_rate", "float", default=1.0, allow_missing_column=True),
    ColumnSpec("home_map", "str", default="", allow_missing_column=True),
    ColumnSpec("home_x", "int", default=0, allow_missing_column=True),
    ColumnSpec("home_y", "int", default=0, allow_missing_column=True),
    ColumnSpec("pk_village_map", "str", default="", allow_missing_column=True),
    ColumnSpec("pk_village_x", "int", default=0, allow_missing_column=True),
    ColumnSpec("pk_village_y", "int", default=0, allow_missing_column=True),
    ColumnSpec("fixes_json", "json", default=[], allow_missing_column=True),
    ColumnSpec("safe_zones_json", "json", default=[], allow_missing_column=True),
    ColumnSpec("quest_requirements_json", "json", default=[], allow_missing_column=True),
)

GATE_COLUMNS = (
    ColumnSpec("gate_id", "int", required=True, default=0),
    ColumnSpec("source_map", "str", required=True, default=""),
    ColumnSpec("source_x", "int", required=True, default=0),
    ColumnSpec("source_y", "int", required=True, default=0),
    ColumnSpec("target_map", "str", required=True, default=""),
    ColumnSpec("target_x", "int", required=True, default=0),
    ColumnSpec("target_y", "int", required=True, default=0),
    ColumnSpec("require_item", "bool", required=True, default=False),
    ColumnSpec("required_item_id", "int", required=True, default=0),
)

DROP_COLUMNS = (
    ColumnSpec("monster_template_id", "int", required=True, default=0),
    ColumnSpec("item_id", "int", required=True, default=0),
    ColumnSpec("drop_rate", "float", required=True, default=0.0),
    ColumnSpec("min_count", "int", required=True, default=0),
    ColumnSpec("max_count", "int", required=True, default=0),
    ColumnSpec("rarity", "int", required=True, default=1),
    ColumnSpec("boss_bonus", "float", required=True, default=0.0),
)

SHOP_COLUMNS = (
    ColumnSpec("store_id", "int", required=True, default=0),
    ColumnSpec("name", "str", required=True, default=""),
    ColumnSpec("buy_rate", "float", required=True, default=1.0),
    ColumnSpec("sell_rate", "float", required=True, default=0.5),
    ColumnSpec("items_json", "json", required=True, default=[]),
)

MONSTER_SPAWN_COLUMNS = (
    ColumnSpec("spawn_id", "int", required=True, default=0),
    ColumnSpec("map_id", "int", required=True, default=0),
    ColumnSpec("center_x", "int", required=True, default=0),
    ColumnSpec("center_y", "int", required=True, default=0),
    ColumnSpec("spawn_radius", "int", required=True, default=0),
    ColumnSpec("monster_template_id", "int", required=True, default=0),
    ColumnSpec("patrol_radius", "int", required=True, default=0),
    ColumnSpec("respawn_interval", "float", required=True, default=0.0),
    ColumnSpec("max_count", "int", required=True, default=0),
    ColumnSpec("aggro_range", "int", required=True, default=0),
    ColumnSpec("attack_range", "int", required=True, default=0),
)

NPC_COLUMNS = (
    ColumnSpec("npc_id", "int", required=True, default=0),
    ColumnSpec("template_id", "int", required=True, default=0),
    ColumnSpec("name", "str", required=True, default=""),
    ColumnSpec("type", "str", required=True, default=""),
    ColumnSpec("map_id", "int", required=True, default=0),
    ColumnSpec("x", "int", required=True, default=0),
    ColumnSpec("y", "int", required=True, default=0),
    ColumnSpec("direction", "int", required=True, default=0),
    ColumnSpec("enabled", "bool", required=True, default=True),
    ColumnSpec("store_id", "int", required=True, default=0),
)


TABLE_REGISTRY = {
    "items": TableSpec(
        name="items",
        workbook="items.xlsx",
        sheet="items",
        artifact_file="items.json",
        payload_root="items",
        primary_key="id",
        columns=ITEM_COLUMNS,
        references=(),
        exporter=_export_items,
    ),
    "skills": TableSpec(
        name="skills",
        workbook="skills.xlsx",
        sheet="skills",
        artifact_file="skills.json",
        payload_root="skills",
        primary_key="id",
        columns=SKILL_COLUMNS,
        references=(),
        exporter=_export_skills,
    ),
    "maps": TableSpec(
        name="maps",
        workbook="maps.xlsx",
        sheet="maps",
        artifact_file="maps.json",
        payload_root="maps",
        primary_key="map_id",
        columns=MAP_COLUMNS,
        references=(),
        exporter=_export_maps,
    ),
    "gates": TableSpec(
        name="gates",
        workbook="gates.xlsx",
        sheet="gates",
        artifact_file="gates.json",
        payload_root="gates",
        primary_key="gate_id",
        columns=GATE_COLUMNS,
        references=(
            ReferenceSpec(
                source_label="gates.source_map",
                target_table="maps",
                target_field="map_id",
                extractor=lambda row: [row["source_map"]],
                value_transform=str,
                target_transform=str,
            ),
            ReferenceSpec(
                source_label="gates.target_map",
                target_table="maps",
                target_field="map_id",
                extractor=lambda row: [row["target_map"]],
                value_transform=str,
                target_transform=str,
            ),
            ReferenceSpec(
                source_label="gates.required_item_id",
                target_table="items",
                target_field="id",
                extractor=lambda row: [row["required_item_id"]] if row["require_item"] else [],
            ),
        ),
        exporter=_export_gates,
    ),
    "drops": TableSpec(
        name="drops",
        workbook="drops.xlsx",
        sheet="drops",
        artifact_file="drops.json",
        payload_root="drop_tables",
        primary_key="monster_template_id",
        columns=DROP_COLUMNS,
        references=(
            ReferenceSpec(
                source_label="drops.item_id",
                target_table="items",
                target_field="id",
                extractor=lambda row: [item["item_id"] for item in row["items"]],
            ),
        ),
        exporter=_export_drops,
    ),
    "shops": TableSpec(
        name="shops",
        workbook="shops.xlsx",
        sheet="shops",
        artifact_file="shops.json",
        payload_root="shops",
        primary_key="store_id",
        columns=SHOP_COLUMNS,
        references=(
            ReferenceSpec(
                source_label="shops.items.item_id",
                target_table="items",
                target_field="id",
                extractor=lambda row: [item["item_id"] for item in row["items"]],
            ),
        ),
        exporter=_export_shops,
    ),
    "monster_spawns": TableSpec(
        name="monster_spawns",
        workbook="monster_spawns.xlsx",
        sheet="monster_spawns",
        artifact_file="monster_spawns.json",
        payload_root="spawn_points",
        primary_key="spawn_id",
        columns=MONSTER_SPAWN_COLUMNS,
        references=(
            ReferenceSpec(
                source_label="monster_spawns.map_id",
                target_table="maps",
                target_field="map_id",
                extractor=lambda row: [row["map_id"]],
            ),
        ),
        exporter=_export_monster_spawns,
    ),
    "npcs": TableSpec(
        name="npcs",
        workbook="npcs.xlsx",
        sheet="npcs",
        artifact_file="npcs.json",
        payload_root="npcs",
        primary_key="npc_id",
        columns=NPC_COLUMNS,
        references=(
            ReferenceSpec(
                source_label="npcs.map_id",
                target_table="maps",
                target_field="map_id",
                extractor=lambda row: [row["map_id"]],
            ),
            ReferenceSpec(
                source_label="npcs.store_id",
                target_table="shops",
                target_field="store_id",
                extractor=lambda row: [row["store_id"]],
            ),
        ),
        exporter=_export_npcs,
    ),
}


def _resolve_tables(requested_tables: list[str]) -> list[str]:
    resolved = list(requested_tables)
    seen = set(requested_tables)
    visited = set()

    def visit(table_name: str):
        if table_name in visited:
            return
        visited.add(table_name)
        for reference in TABLE_REGISTRY[table_name].references:
            if reference.target_table not in seen:
                seen.add(reference.target_table)
                resolved.append(reference.target_table)
            visit(reference.target_table)

    for table_name in requested_tables:
        visit(table_name)
    return resolved


def _row_identifier(spec: TableSpec, row: dict[str, Any]) -> str:
    return f"{spec.name}[{spec.primary_key}={row[spec.primary_key]}]"


def _validate_references(exports: dict[str, TableExport]):
    target_sets = {}
    for table_name, exported in exports.items():
        target_sets[table_name] = {}
        for column in exported.records:
            pass
        for field_name in {reference.target_field for reference in exported.spec.references}:
            (void := field_name)

    for table_name, exported in exports.items():
        spec = exported.spec
        for reference in spec.references:
            target_records = exports[reference.target_table].records
            target_transform = reference.target_transform or _identity
            target_values = {
                target_transform(target_record[reference.target_field])
                for target_record in target_records
            }

            value_transform = reference.value_transform or _identity
            for record in exported.records:
                values = reference.extractor(record)
                for value in values:
                    normalized_value = value_transform(value)
                    if reference.nullable and normalized_value in {None, ""}:
                        continue
                    if normalized_value not in target_values:
                        raise ValueError(
                            f"{_row_identifier(spec, record)} {reference.source_label} "
                            f"references missing {reference.target_table}.{reference.target_field}={value}"
                        )


def _write_outputs(exports: dict[str, TableExport], requested_tables: list[str], out_dir: Path, generated_at: str):
    out_dir.mkdir(parents=True, exist_ok=True)

    artifacts = []
    for table_name in sorted(requested_tables):
        exported = exports[table_name]
        payload_text = json.dumps(exported.payload(), ensure_ascii=False, indent=2) + "\n"
        artifact_path = out_dir / exported.spec.artifact_file
        artifact_path.write_text(payload_text, encoding="utf-8")
        artifacts.append(
            {
                "name": table_name,
                "file": exported.spec.artifact_file,
                "hash": hashlib.sha256(payload_text.encode("utf-8")).hexdigest(),
                "row_count": len(exported.records),
            }
        )

    manifest_payload = {
        "bundle_type": BUNDLE_TYPE,
        "schema_version": SCHEMA_VERSION,
        "generated_at": generated_at,
        "artifacts": artifacts,
    }
    manifest_text = json.dumps(manifest_payload, ensure_ascii=False, indent=2) + "\n"
    (out_dir / "manifest.json").write_text(manifest_text, encoding="utf-8")


def parse_args(argv):
    parser = argparse.ArgumentParser(description="Export gameplay runtime config artifacts")
    parser.add_argument("--source-dir", required=True)
    parser.add_argument("--out-dir", required=True)
    parser.add_argument("--tables", nargs="+")
    parser.add_argument("--generated-at", default="")
    parser.add_argument("--exported-at", default="", help=argparse.SUPPRESS)
    parser.add_argument("--generation", default="", help=argparse.SUPPRESS)
    return parser.parse_args(argv)


def main(argv=None):
    args = parse_args(argv if argv is not None else sys.argv[1:])
    requested_tables = args.tables or list(TABLE_REGISTRY.keys())
    requested_tables = list(dict.fromkeys(requested_tables))
    unsupported = sorted(table for table in requested_tables if table not in TABLE_REGISTRY)
    if unsupported:
        raise ValueError("unsupported tables: " + ", ".join(unsupported))

    generated_at = args.generated_at or args.exported_at
    if not generated_at:
        generated_at = datetime.now(timezone.utc).isoformat()

    source_dir = Path(args.source_dir)
    load_order = _resolve_tables(requested_tables)
    exports = {}
    for table_name in load_order:
        spec = TABLE_REGISTRY[table_name]
        exports[table_name] = spec.exporter(source_dir, spec)

    _validate_references(exports)
    _write_outputs(exports, requested_tables, Path(args.out_dir), generated_at)
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except ValueError as exc:
        print(str(exc), file=sys.stderr)
        raise SystemExit(1)
